from .forms import EmployeeForm, PositionForm, DepartmentForm, RankForm, PositionTitleForm
from .models import Employee, Rank, Position, PositionTitle, Career, Education, Certification, SalaryInfo, Department
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q
from datetime import datetime, date
import os
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from django.contrib import messages

def home(request):
    """
    메인 홈 화면을 렌더링합니다.
    """
    return render(request, 'myapp/home.html')

def department_list(request):
    departments = Department.objects.all().order_by('code')
    form = DepartmentForm()

    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('department_list')

    return render(request, 'myapp/department_list.html', {
        'departments': departments,
        'form': form,
    })

def delete_department(request, pk):
    department = get_object_or_404(Department, pk=pk)

    # 사원이 해당 부서를 참조하고 있다면 삭제 제한
    if department.employee_set.exists():
        # 삭제 방지: 메시지 기능이 있다면 전달해도 좋음
        return redirect('department_list')

    department.delete()
    return redirect('department_list')

def edit_department(request, pk):
    department = get_object_or_404(Department, pk=pk)

    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            form.save()
            return redirect('department_list')
    else:
        form = DepartmentForm(instance=department)

    return render(request, 'myapp/department_edit.html', {
        'form': form,
        'department': department,
    })

def register_employee(request, pk=None):
    """
    사원 등록 및 수정 기능을 처리합니다.
    
    - pk가 제공되면 해당 사원의 정보를 수정합니다.
    - pk가 없으면 새로운 사원 등록 폼을 제공합니다.
    - 등록 또는 수정 후에는 다시 등록 페이지로 리디렉션합니다.
    - 좌측에 전체 사원 리스트를 함께 표시합니다.
    """
    if pk:
        employee = get_object_or_404(Employee, pk=pk)
    else:
        employee = None

    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES, instance=employee)
        if form.is_valid():
            employee = form.save(commit=False)  # 저장은 잠시 보류
            
            # 주민등록번호 기반 age 수동 재계산
            ssn = form.cleaned_data.get('ssn')
            if ssn and len(ssn) >= 7:
                yy = int(ssn[:2])
                mm = int(ssn[2:4])
                dd = int(ssn[4:6])
                code = ssn[6]
                year = 1900 + yy if code in ['1', '2'] else 2000 + yy
                birth = date(year, mm, dd)
                today = date.today()
                age = today.year - birth.year - ((today.month, today.day) < (birth.month, birth.day))
                employee.age = age

            employee.save()
            
            # ✅ 경력 저장
            Career.objects.filter(employee=employee).delete()  # 수정 시 초기화
            companies = request.POST.getlist('career_company[]')
            departments = request.POST.getlist('career_department[]')
            ranks = request.POST.getlist('career_rank[]')
            positions = request.POST.getlist('career_position[]')
            starts = request.POST.getlist('career_start[]')
            ends = request.POST.getlist('career_end[]')
            reasons = request.POST.getlist('career_reason[]')
            for i in range(len(companies)):
                if companies[i]:
                    Career.objects.create(
                        employee=employee,
                        company=companies[i],
                        department=departments[i],
                        rank=ranks[i],
                        position=positions[i],
                        start_date=starts[i],
                        end_date=ends[i],
                        reason=reasons[i],
                    )

            # ✅ 학력 저장
            Education.objects.filter(employee=employee).delete()
            schools = request.POST.getlist('edu_school[]')
            majors = request.POST.getlist('edu_major[]')
            edu_starts = request.POST.getlist('edu_start[]')
            edu_ends = request.POST.getlist('edu_end[]')
            for i in range(len(schools)):
                if schools[i]:
                    Education.objects.create(
                        employee=employee,
                        school=schools[i],
                        major=majors[i],
                        start_date=edu_starts[i],
                        end_date=edu_ends[i],
                    )

            # ✅ 자격 저장
            Certification.objects.filter(employee=employee).delete()
            cert_names = request.POST.getlist('cert_name[]')
            cert_levels = request.POST.getlist('cert_level[]')
            cert_acqs = request.POST.getlist('cert_acquired[]')
            cert_exps = request.POST.getlist('cert_expiry[]')
            for i in range(len(cert_names)):
                if cert_names[i]:
                    Certification.objects.create(
                        employee=employee,
                        name=cert_names[i],
                        level=cert_levels[i],
                        acquired_date=cert_acqs[i],
                        expiry_date=cert_exps[i],
                    )

            # ✅ 연봉 및 계좌정보 저장 (OneToOne)
            SalaryInfo.objects.update_or_create(
                employee=employee,
                defaults={
                    'annual_salary': request.POST.get('annual_salary') or 0,
                    'monthly_salary': request.POST.get('monthly_salary') or 0,
                    'bank_code': request.POST.get('bank_code'),
                    'bank_name': request.POST.get('bank_name'),
                    'account_number': request.POST.get('account_number'),
                    'account_holder': request.POST.get('account_holder'),
                    # ✅ 급여항목 저장
                    'base_salary': request.POST.get('base_salary') or 0,
                    'fixed_overtime_pay': request.POST.get('fixed_overtime_pay') or 0,
                    'meal_allowance': request.POST.get('meal_allowance') or 0,
                    'duty_allowance': request.POST.get('duty_allowance') or 0,
                    'commute_allowance': request.POST.get('commute_allowance') or 0,
                    'childcare_allowance': request.POST.get('childcare_allowance') or 0,
                }
            )

            # ✅ 퇴사일에 따라 상태 자동 설정
            if form.is_valid():
                employee = form.save(commit=False)

                # ✅ 휴직인 경우는 유지
                if employee.status != '휴직':
                    if employee.resign_date:
                        employee.status = '퇴직'
                    else:
                        employee.status = '재직'

                employee.save()
                return redirect('register')

    else:
        form = EmployeeForm(instance=employee)
        career_list = employee.careers.all() if employee else []
        education_list = employee.educations.all() if employee else []
        cert_list = employee.certifications.all() if employee else []
        salary_info = getattr(employee, 'salary_info', None)

        bank_map = {
        '004': '국민은행',
        '011': '농협은행',
        '020': '우리은행',
        '088': '신한은행',
        '081': '하나은행',
        '003': '기업은행',
        '023': 'SC제일은행',
        '027': '한국씨티은행',
        '039': '경남은행',
        '034': '광주은행',
    }
    
    # ✅ 검색어 처리
    search_query = request.GET.get('search', '')  # URL에 ?search= 입력 시 감지

    if search_query:
        employees = Employee.objects.filter(
            Q(name__icontains=search_query) | Q(employee_id__icontains=search_query)
        ).order_by('employee_id')
    else:
        employees = Employee.objects.all().order_by('employee_id')
    
    return render(request, 'myapp/register.html', {
    'form': form,
    'employees': employees,
    'employee': form.instance,  # ✅ 템플릿에서 employee.photo 사용 가능
    'career_list': career_list,
    'education_list': education_list,
    'cert_list': cert_list,
    'salary_info': salary_info,
    'bank_map': bank_map.items(),
})

def employee_list(request):
    """
    사원 목록을 필터링하여 표시합니다.

    - 'active' 필터: 재직자만 조회 (resign_date가 없는 경우)
    - 'retired' 필터: 퇴직자만 조회 (resign_date가 있는 경우)
    - 필터가 없으면 전체 사원을 조회합니다.
    """
    filter_option = request.GET.get('filter')
    search_query = request.GET.get('search')  # 🔥 검색어 받기
    department_query = request.GET.get('department')  # 🔥 부서 검색어 받기
    hire_date_from = request.GET.get('hire_date_from')
    hire_date_to = request.GET.get('hire_date_to')
    resign_date_from = request.GET.get('resign_date_from')
    resign_date_to = request.GET.get('resign_date_to')
    
    if filter_option == 'active':  # 재직자
        employees = Employee.objects.filter(resign_date__isnull=True).order_by('employee_id')
    elif filter_option == 'retired':  # 퇴직자
        employees = Employee.objects.filter(resign_date__isnull=False).order_by('employee_id')
    else:  # 전체
        employees = Employee.objects.all().order_by('employee_id')
        
    # 🔥 검색어가 있으면 이름 또는 사번에서 필터
    if search_query:
        employees = employees.filter(
            Q(name__icontains=search_query) |
            Q(employee_id__icontains=search_query)
        )
       
    # ✅ 입사일 범위 필터
    if hire_date_from:
        employees = employees.filter(hire_date__gte=hire_date_from)
    if hire_date_to:
        employees = employees.filter(hire_date__lte=hire_date_to)

    # ✅ 퇴사일 범위 필터
    if resign_date_from:
        employees = employees.filter(resign_date__gte=resign_date_from)
    if resign_date_to:
        employees = employees.filter(resign_date__lte=resign_date_to)

    employees = employees.order_by('employee_id')

    department_list = Employee.objects.values_list('department', flat=True).distinct().order_by('department')

        # 📌 기준일 필터링
    reference_date_str = request.GET.get('reference_date')
    if reference_date_str:
        reference_date = parse_date(reference_date_str)
        if reference_date:
            employees = employees.filter(
                hire_date__lte=reference_date
            ).filter(
                Q(resign_date__isnull=True) | Q(resign_date__gt=reference_date)
            )

    # 💡 기존의 부서, 상태 등 다른 필터도 여기에 추가 가능

    department_list = Employee.objects.values_list('department', flat=True).distinct()


    return render(request, 'myapp/employee_list.html', {
        'employees': employees,
        'search_query': search_query,
        'department_query': department_query,
        'department_list': department_list,
        'hire_date_from': hire_date_from,
        'hire_date_to': hire_date_to,
        'resign_date_from': resign_date_from,
        'resign_date_to': resign_date_to,
        'reference_date': reference_date_str,
    })
    
def delete_employee(request, pk):
    """
    특정 사원 데이터를 삭제합니다.

    - GET 요청: 삭제 확인 페이지를 렌더링합니다.
    - POST 요청: 삭제를 실행하고 사원 목록 페이지로 리디렉션합니다.
    """
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        employee.delete()
        return redirect('employee_list')
    return render(request, 'myapp/delete_confirm.html', {'employee': employee})

def delete_photo(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if employee.photo:
        # 실제 파일도 삭제
        if os.path.isfile(employee.photo.path):
            os.remove(employee.photo.path)
        employee.photo = None
        employee.save()
    return redirect('register_with_pk', pk=pk)

def export_employees_excel(request):
    employees = Employee.objects.all().order_by('employee_id')

    # ✅ 필터 조건 반영 (조회 페이지와 동일하게 복사)
    filter_option = request.GET.get('filter')
    if filter_option == 'active':
        employees = employees.filter(resign_date__isnull=True)
    elif filter_option == 'retired':
        employees = employees.exclude(resign_date__isnull=True)

    department_query = request.GET.get('department')
    if department_query:
        employees = employees.filter(department=department_query)

    search = request.GET.get('search')
    if search:
        employees = employees.filter(
            Q(name__icontains=search) | Q(employee_id__icontains=search)
        )

    reference_date_str = request.GET.get('reference_date')
    if reference_date_str:
        reference_date = parse_date(reference_date_str)
        if reference_date:
            employees = employees.filter(
                hire_date__lte=reference_date
            ).filter(
                Q(resign_date__isnull=True) | Q(resign_date__gt=reference_date)
            )

    # ✅ 엑셀 작성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "사원명부"


    # ✅ 컬럼명
    columns = [
        "사번", "이름", "성별", "나이", "고용형태", "주민등록번호", "연락처", "부서",
        "직책", "직급", "년차", "직무", "입사일", "퇴사일", "상태", "이메일", "주소"
    ]
    ws.append(columns)

    # ✅ 데이터 행
    for emp in employees:
        ws.append([
            emp.employee_id,
            emp.name,
            emp.gender,
            emp.age,
            emp.employment_type,
            emp.ssn,
            emp.phone,
            emp.department,
            emp.position_title,
            emp.position,
            emp.years_worked,
            emp.role,
            emp.hire_date.strftime("%Y-%m-%d") if emp.hire_date else "",
            emp.resign_date.strftime("%Y-%m-%d") if emp.resign_date else "",
            emp.status,
            emp.email,
            emp.address
        ])

    # ✅ 열 너비 자동 조정
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="사원명부_필터적용.xlsx"'

    wb.save(response)
    return response

def rank_list(request):
    ranks = Rank.objects.all()
    form = RankForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('rank_list')
    return render(request, 'myapp/rank_list.html', {'form': form, 'ranks': ranks})


def edit_rank(request, pk):
    rank = get_object_or_404(Rank, pk=pk)
    form = RankForm(request.POST or None, instance=rank)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('rank_list')
    return render(request, 'myapp/edit_rank.html', {'form': form})


def delete_rank(request, pk):
    rank = get_object_or_404(Rank, pk=pk)
    rank.delete()
    return redirect('rank_list')

def position_list(request):
    positions = PositionTitle.objects.all()
    form = PositionTitleForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('position_list')
    return render(request, 'myapp/position_list.html', {
        'positions': positions,
        'form': form,
    })
    
def edit_position(request, pk):
    position = get_object_or_404(PositionTitle, pk=pk)
    form = PositionTitleForm(request.POST or None, instance=position)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('position_list')
    return render(request, 'myapp/edit_position.html', {'form': form})


def delete_position(request, pk):
    position = get_object_or_404(PositionTitle, pk=pk)
    position.delete()
    return redirect('position_list')

def position_master_list(request):
    positions = Position.objects.all()
    form = PositionForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('position_master_list')
    return render(request, 'myapp/position_master_list.html', {
        'positions': positions,
        'form': form,
    })


def edit_position_master(request, pk):
    position = get_object_or_404(Position, pk=pk)
    form = PositionForm(request.POST or None, instance=position)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('position_master_list')
    return render(request, 'myapp/edit_position_master.html', {'form': form})


def delete_position_master(request, pk):
    position = get_object_or_404(Position, pk=pk)
    position.delete()
    return redirect('position_master_list')